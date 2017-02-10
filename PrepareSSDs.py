# coding: utf-8
#!/usr/bin/env python

import os
import pync
import csv
from openpyxl import load_workbook

# assigns years to column names in the spreadsheet
years = {'2005': 'F', '2010': 'G', '2020': 'H', '2030': 'I', '2040': 'J', '2050': 'K', '2060': 'L', '2070': 'M', '2080': 'N', '2090': 'O', '2100': 'P'}

# header for the output CSVs
header = '"Major area, region, country or area",Country Code,ISO,2010,2020,2030,2040,2050,2060,2070,2080,2090,2100\n '

# will store the output as one string per scenario and model; each will be written to an individual file at the end
output = {}

countryCodes = {}

# some utility functions

def readCountryCodes():
    input_file = csv.DictReader(open("CountryCodes.csv"), delimiter=';')
    for row in input_file:
        countryCodes[row['ISO_A3']] = {'UN_A3': int(row['UN_A3']), 'Name': row['NAME'] }

# gets the number for a given year and row from a spreadsheet in a workbook
def getIntPop(sheet, row, year):
    return (int((sheet[str(years[str(year)]) + str(row)].value) * 1000000))

def getUrbRate(sheet, row, year):
    return sheet[str(years[str(year)]) + str(row)].value



# adds the line to the output dict, either as a new entry, or, if the combination of type+scenario already exists, attaches the line to that one
def attachToOutput(type, scenario, line):
    key  = type+'-'+scenario

    if key in output:
        output[key] = output[key]+'\n '+line
    else:
        output[key] = header+line

# looks up the numeric country code based on 3 letter country code
def getCountryCode(code):
    return countryCodes[code]['UN_A3']

def getName(code):
    return countryCodes[code]['Name']


# fetches a data row by type (pop or urb) from a sheet
# and returns it as a string that can be passed to attachToOutput
def fetchrow(type, sheet, row):
    code = sheet['C'+str(row)].value
    out = str(getCountryCode(code))+','+code+',"'+getName(code)+'"'
    for y in range(2010, 2101, 10):
        if type == 'pop':
            out = out + ',' + str(getIntPop(sheet, row, y))
        if type == 'urb':
            out = out + ',' + str(getUrbRate(sheet, row, y))

    return out


# saves the output dict as a series of CSVs
def saveOutput():
    for f in output:
        file = open(f+'.csv','w')
        file.write(output[f])
        file.close()


# here we go 🚀

os.chdir('/Users/carsten/Dropbox/CISC Data/SSPs')

readCountryCodes()

# load Excel workbooks:
pop = load_workbook('SSPs_Population_Countries.xlsx')
urb = load_workbook('SSPs_Urbanization_Countries.xlsx')

# Fetch the sheets that we need
p = pop['All']
u = urb['All']

# loop over them to pull out the values that we need
for row in range(2,4000):
    if p['A'+str(row)].value == 'NCAR':
        scenario = p['B'+str(row)].value
        attachToOutput('pop', scenario, fetchrow('pop', p, row))

    if u['A'+str(row)].value == 'NCAR':
        scenario = u['B'+str(row)].value
        attachToOutput('urbRate', scenario, fetchrow('urb', u, row))

saveOutput()

pync.Notifier.notify('Preparing SSDs complete 🎉', title='Population projections')
