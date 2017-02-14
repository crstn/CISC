# coding: utf-8
#!/usr/bin/env python

import os
import pync
import csv
from openpyxl import load_workbook

# assigns years to column names in the spreadsheet
years = {'2005': 'F', '2010': 'G', '2020': 'H', '2030': 'I', '2040': 'J',
         '2050': 'K', '2060': 'L', '2070': 'M', '2080': 'N', '2090': 'O',
         '2100': 'P'}

# header for the output CSVs
header = 'Country code,ISO,2010,2020,2030,2040,2050,2060,2070,2080,2090,2100,"Major area, region, country or area"\n'

# will store the output as one string per scenario and model; each will be
# written to an individual file at the end
output = {}

countryCodes = {}

# some utility functions


def readCountryCodes():
    input_file = csv.DictReader(open("CountryCodes.csv"), delimiter=';')
    for row in input_file:
        countryCodes[row['ISO_A3']] = {
            'UN_A3': int(row['UN_A3']), 'Name': row['NAME']}

# gets the number for a given year and row from a spreadsheet in a workbook


def getIntPop(sheet, row, year):
    return (int((sheet[str(years[str(year)]) + str(row)].value) * 1000000))


def getUrbRate(sheet, row, year):
    return sheet[str(years[str(year)]) + str(row)].value


# adds the line to the output dict, either as a new entry, or, if the
# combination of type+scenario already exists, attaches the line to that
# one
def attachToOutput(type, scenario, line):
    key = type + '-' + scenario

    if key in output:
        output[key] = output[key] + '\n' + line
    else:
        output[key] = header + line

# looks up the numeric country code based on 3 letter country code


def getCountryCode(code):
    return countryCodes[code]['UN_A3']


def getName(code):
    return countryCodes[code]['Name']


# fetches a data row by type (pop or urb) from a sheet
# and returns it as a string that can be passed to attachToOutput
def fetchrow(type, sheet, row):
    code = sheet['C' + str(row)].value
    out = str(getCountryCode(code)) + ',' + code
    for y in range(2010, 2101, 10):
        if type == 'pop':
            out = out + ',' + str(getIntPop(sheet, row, y))
        if type == 'urb':
            out = out + ',' + str(getUrbRate(sheet, row, y))

    return out + ',"' + getName(code) + '"'


# saves the output dict as a series of CSVs
def saveOutput():
    for f in output:
        file = open(f + '.csv', 'w')
        file.write(output[f])
        file.close()


# here we go 🚀

os.chdir('/Users/carsten/Dropbox/CISC Data/SSPs')

readCountryCodes()

# load Excel workbooks:
pop = load_workbook('SSPs_Population_Countries.xlsx')
urb = load_workbook('SSPs_Urbanization_Countries.xlsx')

# Fetch the sheets that we need
p = pop['All']
u = urb['All']

# loop over them to pull out the values that we need
for row in range(2, 4000):
    if p['A' + str(row)].value == 'NCAR':
        scenario = p['B' + str(row)].value
        attachToOutput('pop', scenario, fetchrow('pop', p, row))

    if u['A' + str(row)].value == 'NCAR':
        scenario = u['B' + str(row)].value
        attachToOutput('urbRate', scenario, fetchrow('urb', u, row))



# calculate actual urban population from urbanization rates:
for ssp in range(1,6): # SSP1–5
    pn = output['pop-SSP'+str(ssp)]
    ur = output['urbRate-SSP'+str(ssp)]

    # for each row in pn, find the matching ur:
    poplines = pn.split('\n')
    urblines = ur.split('\n')

    if len(poplines) != len(urblines):
        print "Population file and urbanization have different sizes"
    else:
        for i in range(1,len(poplines)): # skip header
            plparts = poplines[i].split(',')
            ulparts = urblines[i].split(',')

            if len(plparts) != len(ulparts):
                print "Different number of parts after comma splitting"

            else:

                for j in range(0,len(plparts)):

                    if j == 0:
                        if plparts[j] != ulparts[j]:
                            print "Different country codes"
                        outputline = plparts[j]

                    # columns 2–11 are the numbers we want to process:
                    elif j in range(2,12):
                        # compute urban pop from urb rate and total pop
                        urbnum = int(float(plparts[j]) * float(ulparts[j]) / 100.0)
                        outputline = outputline + ',' + str(urbnum)
                    else: # everything else is just copied over to the output
                        outputline = outputline + ',' + plparts[j]

                attachToOutput('urbpop', 'SSP'+str(ssp), outputline)

saveOutput()

pync.Notifier.notify('Preparing SSDs complete 🎉',
                     title='Population projections')
